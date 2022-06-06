import base64
import graphene
from graphene import ID, List, Boolean, String, Date
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from ..utils.enum.time_periods import Period, UnitOfTime
from ..utils.openpyxl import copy_sheet_to_workbook, create_empty_workbook
from ....xlsx.constructor_report_xlsx import InfoTitle, Flag
from ....xlsx.moduls.unloading_order_xlsx import generate_file
from ..utils.auth.decorators import login_or_permissions_required
from ..queries.search_packages_info import prepare_packages_info


import calendar as calendar

calendar = calendar.Calendar()


package_field_title = (
    InfoTitle('Идентификатор пакета', 'id', Flag.key.name),
    InfoTitle('Наименование пакета', 'title', Flag.key.name),
    InfoTitle('Название города', 'city__title', Flag.key.name),
    InfoTitle('Год', 'year', Flag.key.name),
    InfoTitle('Месяц', 'month', Flag.key.name),
)

construction_side_field_title = (
    InfoTitle('Идентификатор стороны', 'id', Flag.key.name),
    InfoTitle('Пакет', 'package_id', Flag.key.name),
    InfoTitle('Адрес', 'construction__marketing_address__address', Flag.key.name),
    InfoTitle('Рекламная сторона', 'advertising_side__title', Flag.key.name),
    InfoTitle('Город', 'package__city__title', Flag.key.name),
    InfoTitle('Доступность', lambda v: 'Да' if v['availability_side'] else 'Нет', Flag.func.name),
    InfoTitle('В архиве', lambda v: 'Да' if v['is_archive'] else 'Нет', Flag.func.name),
)


def extend_packages_by_reserved_sides_stat(workbook, packages_sides_counts, package_reserved_sides_by_period_counts):
    sheet = workbook.get_sheet_by_name('Пакеты')
    packages_counts = {}
    for count in packages_sides_counts:
        packages_counts[count['package_id']] = count['count']

    mr = sheet.max_row
    mc = sheet.max_column

    # Всего сторон в пакете
    sheet.cell(row=1, column=mc + 1).value = 'Всего сторон в пакете'
    for i in range(2, mr + 1):
        package_id = sheet.cell(row=i, column=1).value
        sheet.cell(row=i, column=mc + 1).value = packages_counts[package_id] if package_id in packages_counts else 0
    mc += 1

    # Занятые стороны по пакетам
    offset = 1
    for dataset in package_reserved_sides_by_period_counts:
        sheet.cell(row=1, column=mc + offset).value = dataset[0][0].strftime("%d-%m-%Y") + ' - ' + dataset[0][1].strftime("%d-%m-%Y")
        for i in range(2, mr + 1):
            package_id = sheet.cell(row=i, column=1).value
            sheet.cell(row=i, column=mc + offset).value = dataset[1][package_id] if package_id in dataset[1] else 0
        offset += 1


class DownloadPackagesInfo(graphene.Mutation):
    class Arguments:
        packages_ids = List(
            ID,
            description='Список идентификаторов пакетов', required=True
        )
        unit_of_time = UnitOfTime(
            description='Единица времени, по которой производится агрегация'
        )
        period = Period(
            description='Период, за который производится выборка. Если указан тип CUSTOM, фильтрует в диапазоне [date_from, date_to]'
        )
        date_from = Date(
            description='Дата начала периода. Указанный день включается в выборку [date_from:date_to)'
        )
        date_to = Date(
            description='Дата конца периода. Указанный день не включается в выборку [date_from:date_to)'
        )

    file = String(description='Файл *.xlsx в кодировке Base64')
    ok = Boolean()

    @login_or_permissions_required(login_required=False, permissions=('api.download_packages_info', ))
    def mutate(root, info, **input):
        packages, sides_values, count_sides_by_packages, counts, offset, limit, count = prepare_packages_info(input)

        packages_file = generate_file(packages, 'p_info_xlsx_file', package_field_title)
        sides_file = generate_file(sides_values, 'p_info_xlsx_file', construction_side_field_title)

        packages_ws = load_workbook(packages_file).worksheets[0]
        sides_ws = load_workbook(sides_file).worksheets[0]

        output_wb = create_empty_workbook()

        copy_sheet_to_workbook(output_wb, packages_ws, 'Пакеты')
        copy_sheet_to_workbook(output_wb, sides_ws, 'Стороны')

        extend_packages_by_reserved_sides_stat(output_wb, count_sides_by_packages, counts)

        output = base64.b64encode(save_virtual_workbook(output_wb)).decode('ascii')

        return DownloadPackagesInfo(file=output, ok=True)
