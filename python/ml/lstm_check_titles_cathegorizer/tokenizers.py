import os
from collections import OrderedDict
import string
import re

from pymystem3 import Mystem
from django.core.management.base import BaseCommand, CommandError

# os.environ["CUDA_VISIBLE_DEVICES"] = "-1"  # Force CPU
from .consts import TRAIN_FILE

os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3' # 0 = all messages are logged, 3 - INFO, WARNING, and ERROR messages are not printed

import numpy as np
import pandas as pd
import openpyxl

from keras.models import Sequential
from keras.layers import Dense, LSTM, Embedding, RepeatVector, Dropout
from keras.preprocessing.text import Tokenizer
from keras.preprocessing.sequence import pad_sequences
from keras.models import load_model
from keras import optimizers
from sklearn.model_selection import train_test_split


def populate_cols_idx(sheet, header_row=1):
    dst = OrderedDict()
    for col_idx in range(1, sheet.max_column + 1):
        k = sheet.cell(header_row, col_idx).value
        dst[k.strip() if isinstance(k, str) else k] = col_idx
    return dst

def strip(s):
    return s.strip() if (s is not None and isinstance(s, str)) else s

class SheetValue:
    def __init__(self, sheet, cols):
        self.sheet = sheet
        self.cols = cols
        self.row_idx = 0

    def set_row_idx(self, row_idx):
        self.row_idx = row_idx

    def get(self, key):
        col_idx = self.cols[key]
        x = self.sheet.cell(self.row_idx, col_idx).value
        return strip(str(x)) if x is not None else x

    def get_or_emptystr(self, key):
        col_idx = self.cols[key]
        x = self.sheet.cell(self.row_idx, col_idx).value
        return strip(str(x)) if x is not None else ''


def get_sheet_value(wb_fname, wb_sheet, header_row=1):
    wb = openpyxl.load_workbook(wb_fname)
    sheet = wb.get_sheet_by_name(wb_sheet)
    cols = populate_cols_idx(sheet, header_row)
    sheet_value = SheetValue(sheet, cols)
    print(cols)
    return wb, cols, sheet, sheet_value


TRANSFORM_RE = re.compile(r'''[-:+.,!@#$%^&*()_'"\\<>?/]''')
TRANSFORM_RE2 = re.compile(r'''[ \t]+''')
MYSTEM = Mystem()


def prepare_string(s):
    base_str = TRANSFORM_RE.sub(' ', s)
    base_str = base_str.translate(str.maketrans('', '', string.punctuation)).lower()
    base_str = TRANSFORM_RE2.sub(' ', base_str)
    # base_str = ''.join(MYSTEM.lemmatize(base_str))
    return base_str


def prepare_and_lemmatize_string(s):
    base_str = TRANSFORM_RE.sub(' ', s)
    base_str = base_str.translate(str.maketrans('', '', string.punctuation)).lower()
    base_str = TRANSFORM_RE2.sub(' ', base_str)
    base_str = ''.join(MYSTEM.lemmatize(base_str))
    return base_str


def check_exec_time_mystem_one_text(texts):
    lol = lambda lst, sz: [lst[i:i+sz] for i in range(0, len(lst), sz)]
    txtpart = lol(texts, 1000)
    res = []
    for txtp in txtpart:
        alltexts = ' '.join([txt + ' brrrrr ' for txt in txtp])

        words = MYSTEM.lemmatize(alltexts)
        doc = []
        for txt in words:
            if txt != '\n' and txt.strip() != '':
                if txt == 'brrrrr':
                    res.append(doc)
                    doc = []
                else:
                    doc.append(txt)
    return res


class Tokenizers:
    def __init__(self):
        wb, cols, sheet, sheet_value = get_sheet_value(os.path.join('data', TRAIN_FILE), 'Sheet')
        categories = set()
        names_for_transforming = []

        for row_idx in range(3, sheet.max_row + 1):
            sheet_value.set_row_idx(row_idx)
            cat, tp = (sheet_value.get_or_emptystr('Категория товара'), sheet_value.get_or_emptystr('Тип товара'))
            if cat or tp:
                categories.add((cat, tp))
            names_for_transforming.append(prepare_string(sheet_value.get_or_emptystr('Наименование товара')))

        transformed_names = check_exec_time_mystem_one_text(names_for_transforming)
        assert len(names_for_transforming) == len(transformed_names)

        self.indexed_categories = {cat: i for i, cat in enumerate(sorted(categories))}
        self.reversed_indexed_categories = {i: cat for cat, i in self.indexed_categories.items()}

        # for x in sorted(self.indexed_categories):
        #     print(x, self.indexed_categories[x])
        self.indexed_dataset = {}

        self.dataset = []
        src = []
        for row_idx in range(3, sheet.max_row + 1):
            sheet_value.set_row_idx(row_idx)
            # item = prepare_string(sheet_value.get_or_emptystr('Наименование товара'))
            item = names_for_transforming[row_idx - 3].strip()
            cat, tp = (sheet_value.get_or_emptystr('Категория товара'), sheet_value.get_or_emptystr('Тип товара'))
            if cat or tp:
                key = (cat, tp)
                idx = self.indexed_categories[key]
                self.dataset.append((item, idx))
                self.indexed_dataset[
                    (sheet_value.get_or_emptystr('Данные QR-кода'), sheet_value.get_or_emptystr('Наименование товара'))
                ] = key
                src.append(item)


        self.src_tokenizer = Tokenizer()
        self.src_tokenizer.fit_on_texts(src)
        self.src_vocab_size = len(self.src_tokenizer.word_index) + 1
        self.src_length = 8
        self.dst_vocab_size = len(self.indexed_categories)
        self.dst_length = 1

    # Encode and pad sequences
    def encode_sequences(self, tokenizer, length, lines):
        # integer encode sequences
        seq = tokenizer.texts_to_sequences(lines)
        # pad sequences with 0 values
        seq = pad_sequences(seq, maxlen=length, padding='post')
        return seq
